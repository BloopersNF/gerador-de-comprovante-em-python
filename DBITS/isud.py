import os
from datetime import date
import customtkinter as ctk
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox

def createtxt():
    filename = f"{data} - {company} - {qtd}"
    if os.path.exists(f"documents/{year}/{company}"):
        if os.path.exists(f"documents/{year}/{company}/{filename}.docx"):
            response = messagebox.askyesno("Atenção!" , f"arquivo [{filename}.txt] já existente. Deseja criar mesmo assim?")
            if response == True:
                count = 2
                while os.path.exists(f"documents/{year}/{company}/{filename}({count}).docx"):
                    count = count+1
                file = open(f"documents/{year}/{company}/{filename}({count}).docx", "w")
            else:
                return 0
        else:
            file = open(f"documents/{year}/{company}/{filename}.docx", "w")
    else:
        os.mkdir(f"documents/{year}/{company}")
        if os.path.exists(f"documents/{year}/{company}/{filename}.docx"):
            response = input(f"arquivo [{filename}.txt] já existente.")
            return 0
        file = open(f"documents/{year}/{company}/{filename}.docx", "w")
    source = open("assets/model.txt", "r")
    for row in source.read().replace("{scrap}", str(scrap)).replace("{buyer}", buyer).replace("{company}", company).replace("{qtd}", str(qtd)).replace("{interchangeable}", str(interchangeable)).replace("{seller}", seller):
        file.write(row)
    file.close()
    source.close()

#create a Workbook
sheets = ["Buriti","Finissima","Caninde","Vintani","Lebrinha"]
data = date.today().strftime('%d-%m-%Y')
year = date.today().strftime("%Y")

if not os.path.exists(f"documents/{year}"):
    os.mkdir(f"documents/{year}")

if os.path.isfile(f"documents/{year}/vendas.xlsx"):
    wb = load_workbook(f"documents/{year}/vendas.xlsx")
else:
    wb = Workbook()
    for sheet in sheets:
        wb.create_sheet(sheet)

companys = wb.sheetnames


for sheet in sheets:
    wb[sheet]["A1"] = "Vendedor"
    wb[sheet]["B1"] = "Quantidade"
    wb[sheet]["C1"] = "Data"
    wb[sheet]["D1"] = "intercambiaveis"
    wb[sheet].column_dimensions["A"].widht = 20
    wb[sheet].column_dimensions["C"].widht = 40
    wb.save(f"documents/{year}/vendas.xlsx")

if "Sheet" in companys:
    del wb["Sheet"]

wb.save(f"documents/{year}/vendas.xlsx")
def save():
    global company, qtd,scrap, buyer, seller, interchangeable
    scrap = scrap_entry.get()
    interchangeable = inter_entry.get()
    buyer = buyer_entry.get()
    qtd = qtd_entry.get()
    seller = seller_entry.get()
    company = variable.get()
    count_cells = 1
    while wb[company][f"A{count_cells}"].value != None:
        count_cells = count_cells +1
        print(wb[company][f"A{count_cells}"].value)
    wb[company][f"A{count_cells}"] = seller_entry.get()
    wb[company][f"B{count_cells}"] = qtd_entry.get()
    wb[company][f"C{count_cells}"] = data
    wb[company][f"D{count_cells}"] = scrap_entry.get()
    createtxt()
    wb.save(f"documents/{year}/vendas.xlsx")
    messagebox.showinfo("Sucesso!", "Dados salvos com sucesso.")


ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

#create a window
first = ctk.CTk()
first.title("Dbits")
first.geometry("500x500")
first.resizable(False, False)

first.iconbitmap("assets/robo.ico")

#first tab
seller_label = ctk.CTkLabel(master=first, text="Quem está vendendo:")
seller_label.pack()
seller_entry = ctk.CTkEntry(master=first)
seller_entry.pack()

buyer_label = ctk.CTkLabel(first, text="Quem está comprando:")
buyer_label.pack()
buyer_entry = ctk.CTkEntry(first)
buyer_entry.pack()

qtd_label = ctk.CTkLabel(first, text="Quantidade de garrafões:")
qtd_label.pack()
qtd_entry = ctk.CTkEntry(first)
qtd_entry.pack()

scrap_label = ctk.CTkLabel(first, text="Sucata:")
scrap_label.pack()
scrap_entry = ctk.CTkEntry(first)
scrap_entry.pack()

inter_label = ctk.CTkLabel(first, text="intercambiaveis:")
inter_label.pack()
inter_entry = ctk.CTkEntry(first)
inter_entry.pack()

variable = StringVar()
variable.set("Escolha uma empresa")
company_label = ctk.CTkLabel(first, text="Empresa:")
company_label.pack()
company_menu = ctk.CTkOptionMenu(first, variable=variable,values=companys)
company_menu.pack()

button1 = ctk.CTkButton(first, text="Salvar", command=save)
button1.pack(pady=20)

first.mainloop()
